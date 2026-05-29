from __future__ import annotations

import csv
import json
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.schemas.device import DeviceInput

HEADER_ALIASES = {
    "hostname": "hostname",
    "name": "hostname",
    "ip": "ip_address",
    "ip address": "ip_address",
    "management ip": "ip_address",
    "vendor": "vendor",
    "manufacturer": "vendor",
    "model": "model",
    "site": "site",
    "location": "site",
    "role": "role",
}


def load_inventory_file(path: str | Path) -> list[DeviceInput]:
    source = Path(path)
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return devices_from_records(list(csv.DictReader(handle)))
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        records = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(row)
        ]
        return devices_from_records(records)
    if source.suffix.lower() == ".json":
        payload = json.loads(source.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError("Inventory JSON must contain a list of devices")
        return devices_from_records(payload)
    raise ValueError(f"Unsupported inventory format: {source.suffix}")


def devices_from_records(records: Sequence[dict[str, Any] | DeviceInput]) -> list[DeviceInput]:
    devices: list[DeviceInput] = []
    for record in records:
        if isinstance(record, DeviceInput):
            devices.append(record)
            continue
        normalized = _normalize_record(record)
        if normalized.get("ip_address"):
            devices.append(DeviceInput(**normalized))
    return devices


def _normalize_record(record: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in record.items():
        canonical = HEADER_ALIASES.get(str(key).strip().casefold())
        if canonical:
            normalized[canonical] = str(value or "").strip()
    normalized.setdefault("vendor", "")
    normalized.setdefault("model", "")
    return normalized
