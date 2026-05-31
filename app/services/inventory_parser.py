from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from app.services.device_normalizer import canonicalize_record


class InventoryParser:
    def parse_items(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [canonicalize_record(item) for item in items]

    def parse_json_text(self, text: str) -> list[dict[str, Any]]:
        payload = json.loads(text)
        if not isinstance(payload, list):
            raise ValueError("Inventory JSON must contain a list of objects")
        return self.parse_items([_ensure_mapping(item) for item in payload])

    def parse_csv_text(self, text: str) -> list[dict[str, Any]]:
        rows = list(csv.DictReader(StringIO(text)))
        return self.parse_items([dict(row) for row in rows])

    def parse_xlsx_bytes(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            if not any(row):
                continue
            records.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
        return self.parse_items(records)


def _ensure_mapping(item: Any) -> dict[str, Any]:
    if not isinstance(item, dict):
        raise ValueError("Inventory JSON entries must be objects")
    return item
