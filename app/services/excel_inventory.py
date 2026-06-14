from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "Status",
    "Device Label",
    "Model",
    "IP Address",
    "Vendor",
    "Device Category",
    "Location",
    "Contact",
)


class ExcelInventoryError(ValueError):
    """Raised when the Excel lab inventory cannot be parsed safely."""


@dataclass(frozen=True)
class ExcelInventoryDevice:
    id: str
    status: str | None
    label: str
    hostname: str
    ip_address: str
    vendor: str
    model: str
    category: str | None
    location: str | None
    contact: str | None
    driver_name: str | None
    platform: str | None
    original_vendor: str | None = None
    original_model: str | None = None
    normalized_vendor: str | None = None
    normalized_model: str | None = None


def load_excel_inventory(path: str | Path) -> list[ExcelInventoryDevice]:
    inventory_path = Path(path)
    if not inventory_path.exists():
        raise ExcelInventoryError(f"Excel inventory file not found: {inventory_path}")
    try:
        workbook = load_workbook(inventory_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelInventoryError(f"Unable to read Excel inventory {inventory_path}: {exc}") from exc
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = _next_non_empty_row(rows)
        if headers is None:
            raise ExcelInventoryError("Excel inventory is empty")
        header_map = _header_map(headers)
        missing = [column for column in REQUIRED_COLUMNS if _normalize_header(column) not in header_map]
        if missing:
            raise ExcelInventoryError(f"Excel inventory is missing required column(s): {', '.join(missing)}")
        devices: list[ExcelInventoryDevice] = []
        for row_index, row in enumerate(rows, start=2):
            item = _row_dict(row, header_map)
            if _skip_row(item):
                continue
            devices.append(_device_from_row(item, row_index))
        return devices
    finally:
        workbook.close()


def resolve_excel_device(devices: list[ExcelInventoryDevice], value: str) -> ExcelInventoryDevice:
    needle = _normalize_match(value)
    matches = [
        device
        for device in devices
        if needle
        in {
            _normalize_match(device.id),
            _normalize_match(device.label),
            _normalize_match(device.hostname),
            _normalize_match(device.ip_address),
        }
    ]
    if not matches:
        raise ExcelInventoryError(f"Device {value!r} was not found in Excel inventory")
    if len(matches) > 1:
        labels = ", ".join(f"{device.label} ({device.ip_address})" for device in matches)
        raise ExcelInventoryError(f"Device selector {value!r} matched multiple rows: {labels}")
    return matches[0]


def infer_runtime_hints(vendor: str, model: str, category: str | None = None) -> tuple[str | None, str | None]:
    text = " ".join(part for part in (vendor, model, category or "") if part).casefold()
    if "icmp" in text:
        return "ReadOnlyICMPDriver", "icmp"
    if "unknown product" in text or "unknown snmp product" in text:
        return None, None
    if "securitycode" in text or "continent" in text or "security appliance" in text:
        return "NonSwitchInventoryDriver", "non_switch"
    if "des1100" in text or "des-1100" in text or "unmanaged" in text:
        return "LimitedWebInventoryDriver", "limited_web"
    if "generic" in text:
        return "GenericSSHDriver", "generic"
    if re.search(r"\bqsw[-\s]?(4610|3750)\b", text):
        return "QtechQswDriver", "qtech"
    if re.search(r"\bs(17|23|24|57|67)\d*", text) or re.search(r"\bce68\d*", text) or "vrp" in text:
        return "HuaweiVRPDriver", "vrp"
    if "catalyst" in text or "cat2960" in text or "37xxstack" in text:
        return "CiscoIOSDriver", "ios"
    if any(value in text for value in ("1620", "1820", "1905", "limited web")):
        return "LimitedWebInventoryDriver", "limited_web"
    if "procurve" in text or re.search(r"\b25(10|30)\b", text):
        return "HPEProCurveDriver", "procurve"
    if any(value in text for value in ("comware", "1910", "1920", "1950", "5130", "s4210", "s5500")):
        return "HPComwareDriver", "comware"
    if "mes2324" in text or "mes2348" in text or "mes2448" in text:
        return "EltexMESDriver", "mes"
    if re.search(r"\bbs(2500|6300)\b", text) or "bk-a837" in text:
        return "BulatBSDriver", "bulat"
    if "powerconnect" in text:
        return "DellPowerConnectDriver", "dell"
    return None, None


def _next_non_empty_row(rows: Any) -> tuple[Any, ...] | None:
    for row in rows:
        if any(_cell(value) for value in row):
            return tuple(row)
    return None


def _header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, value in enumerate(headers):
        header = _normalize_header(_cell(value))
        if header:
            mapped[header] = index
    return mapped


def _row_dict(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, str]:
    result: dict[str, str] = {}
    for column in REQUIRED_COLUMNS:
        index = header_map[_normalize_header(column)]
        result[column] = _cell(row[index] if index < len(row) else None)
    return result


def _device_from_row(item: dict[str, str], row_index: int) -> ExcelInventoryDevice:
    label = item["Device Label"]
    ip_address = item["IP Address"]
    original_vendor = item["Vendor"].strip()
    original_model = item["Model"].strip()
    vendor = _normalize_vendor(original_vendor, original_model, item.get("Device Category"))
    model = _normalize_model(original_model)
    if not label or not ip_address or not vendor or not model:
        raise ExcelInventoryError(f"Excel row {row_index} is missing label, IP address, vendor, or model")
    driver_name, platform = infer_runtime_hints(vendor, model, item.get("Device Category"))
    identifier = _stable_id(label=label, ip_address=ip_address, vendor=vendor, model=model)
    return ExcelInventoryDevice(
        id=identifier,
        status=item["Status"] or None,
        label=label,
        hostname=label,
        ip_address=ip_address,
        vendor=vendor,
        model=model,
        category=item["Device Category"] or None,
        location=item["Location"] or None,
        contact=item["Contact"] or None,
        driver_name=driver_name,
        platform=platform,
        original_vendor=original_vendor,
        original_model=original_model,
        normalized_vendor=vendor,
        normalized_model=model,
    )


def _skip_row(item: dict[str, str]) -> bool:
    if not any(item.values()):
        return True
    category = item.get("Device Category", "").strip().casefold()
    status = item.get("Status", "").strip().casefold()
    return category in {"service", "services"} or status == "service"


def _normalize_vendor(value: str, model: str = "", category: str | None = None) -> str:
    text = " ".join(value.split())
    folded = text.casefold()
    combined = f"{folded} {model.casefold()} {(category or '').casefold()}"
    if folded in {"hp", "hpe", "hewlett packard", "hewlett-packard"}:
        return "HPE"
    if folded == "3com":
        return "HPE"
    if folded in {"qtech", "q-tech"} or "qsw" in combined:
        return "QTECH"
    if folded in {"d-link", "dlink"}:
        return "D-Link"
    if "securitycode" in combined or "continent" in combined:
        return "SecurityCode"
    if folded == "huawei":
        return "Huawei"
    if folded == "cisco":
        return "Cisco"
    if folded == "dell":
        return "Dell"
    if folded == "eltex":
        return "Eltex"
    if folded == "bulat":
        return "Bulat"
    return text


def _normalize_model(value: str) -> str:
    text = " ".join(value.replace("_", " ").split())
    folded = text.casefold()
    replacements = {
        "cat2960": "Catalyst 2960",
        "catalyst 37xxstack": "Catalyst 37xxStack",
        "continent 500": "Continent-500",
        "continent-500": "Continent-500",
    }
    if folded in replacements:
        return replacements[folded]
    qsw = re.search(r"\bqsw[-\s]?(4610|3750)\b", folded)
    if qsw:
        return f"QSW-{qsw.group(1)}"
    des = re.search(r"\bdes[-\s]?1100\b", folded)
    if des:
        return "DES1100"
    return text


def _stable_id(*, label: str, ip_address: str, vendor: str, model: str) -> str:
    digest = hashlib.sha256(f"{label}|{ip_address}|{vendor}|{model}".casefold().encode("utf-8")).hexdigest()[:16]
    return f"excel-{digest}"


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _normalize_match(value: str) -> str:
    return value.strip().casefold()


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
