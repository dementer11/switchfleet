from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from .models import Device


HEADER_ALIASES = {
    "device label": "label",
    "label": "label",
    "name": "label",
    "ip address": "ip_address",
    "ip": "ip_address",
    "management ip": "ip_address",
    "vendor": "vendor",
    "manufacturer": "vendor",
    "model": "model",
    "device category": "category",
    "category": "category",
    "location": "location",
    "contact": "contact",
    "status": "status",
}


def load_inventory(path: str | Path) -> list[Device]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return _devices_from_rows(_read_csv(source))
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(source)
    raise ValueError(f"Unsupported inventory format: {source.suffix}")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _read_xlsx(path: Path) -> list[Device]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install with xlsx support: pip install -e .[xlsx]") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    header_index = _find_header_index(rows)
    if header_index is None:
        raise ValueError("Could not find inventory header row with IP Address and Vendor")

    headers = [str(value or "").strip() for value in rows[header_index]]
    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        record = {
            headers[index]: str(value or "").strip()
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(record.values()):
            records.append(record)
    return _devices_from_rows(records)


def _find_header_index(rows: list[list[object]]) -> int | None:
    for index, row in enumerate(rows):
        normalized = {str(value or "").strip().lower() for value in row}
        if "ip address" in normalized and "vendor" in normalized:
            return index
    return None


def _devices_from_rows(rows: Iterable[dict[str, str]]) -> list[Device]:
    devices: list[Device] = []
    for raw in rows:
        normalized = _normalize_record(raw)
        ip = normalized.get("ip_address", "").strip()
        if not ip:
            continue
        devices.append(
            Device(
                label=normalized.get("label", ip),
                ip_address=ip,
                vendor=normalized.get("vendor", ""),
                model=normalized.get("model", ""),
                category=normalized.get("category", ""),
                location=normalized.get("location", ""),
                contact=normalized.get("contact", ""),
                status=normalized.get("status", ""),
                metadata={k: v for k, v in raw.items() if k not in normalized},
            )
        )
    return devices


def _normalize_record(record: dict[str, str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in record.items():
        canonical = HEADER_ALIASES.get(str(key).strip().lower())
        if canonical:
            normalized[canonical] = str(value or "").strip()
    return normalized
